import matplotlib.pyplot as plt
from typing import List, Dict
from server.dataset import Data, DataFilter
from openpyxl import Workbook
import os
from datetime import datetime
import numpy as np

class StatisticGraph:

    CORAL_H = "Coral"
    CORAL_ID_H = "Coral ID"
    NUM_OF_PIXELS_H = "No. of Pixels"
    NUM_OF_HEALTHY_PIXEL_H = "No. of Healthy Pixel"
    NUM_OF_BLEACHED_PIXEL_H = "No. of Bleached Pixel"
    CORAL_COVERAGE_H = "Coral Coverage"
    HEALTHY_COVERAGE_H = "Healthy Coverage"
    BLEACHED_COVERAGE_H = "Bleached Coverage"
    HEALTHY_DISTRIBUTION_H = "Healthy Distribution"
    BLEACHED_DISTRIBUTION_H = "Bleached Distribution"
    NUM_OF_COLONY_H = "No. of Colony"

    def __init__(self, json_item: Dict, label_colors: List[str]):
        self.label_colors = label_colors
        self.json_item = json_item
        self.id_to_name = self.get_id_to_name()
        self.id_to_supercategory = self.get_id_to_supercategory()
        self.name_to_id = self.get_name_to_id()
    
    def extract_excel_data(self) -> Dict:
        data = {}

        image_pixel = self.json_item["images"][0]["width"] * self.json_item["images"][0]["height"]
        for annotation in self.json_item["annotations"]:
            category_id = annotation["category_id"]
            category_name = self.id_to_name[category_id]

            supercategory_name = self.id_to_supercategory[annotation["category_id"]]
            supercategory_id = self.name_to_id[supercategory_name]

            if supercategory_id not in data:
                data[supercategory_id] = {}
                data[supercategory_id][StatisticGraph.CORAL_H] = ""
                data[supercategory_id][StatisticGraph.CORAL_ID_H] = -1
                data[supercategory_id][StatisticGraph.NUM_OF_PIXELS_H] = 0
                data[supercategory_id][StatisticGraph.NUM_OF_HEALTHY_PIXEL_H] = 0
                data[supercategory_id][StatisticGraph.NUM_OF_BLEACHED_PIXEL_H] = 0
                data[supercategory_id][StatisticGraph.NUM_OF_COLONY_H] = 0
                data[supercategory_id][StatisticGraph.CORAL_COVERAGE_H] = 0
                data[supercategory_id][StatisticGraph.HEALTHY_COVERAGE_H] = 0
                data[supercategory_id][StatisticGraph.BLEACHED_COVERAGE_H] = 0
                data[supercategory_id][StatisticGraph.HEALTHY_DISTRIBUTION_H] = 0
                data[supercategory_id][StatisticGraph.BLEACHED_DISTRIBUTION_H] = 0

            data[supercategory_id][StatisticGraph.CORAL_H] = supercategory_name

            if self.is_category_name_is_bleached(category_name):
                data[supercategory_id][StatisticGraph.NUM_OF_BLEACHED_PIXEL_H] += annotation["area"]
            else:
                data[supercategory_id][StatisticGraph.NUM_OF_HEALTHY_PIXEL_H] += annotation["area"]
            data[supercategory_id][StatisticGraph.NUM_OF_PIXELS_H] += annotation["area"]
            data[supercategory_id][StatisticGraph.NUM_OF_COLONY_H] += 1
            data[supercategory_id][StatisticGraph.CORAL_ID_H] = supercategory_id

        for key, value in data.items():
            print(f"key: {key}, value: {value}")
            value[StatisticGraph.CORAL_COVERAGE_H] = value[StatisticGraph.NUM_OF_PIXELS_H] / image_pixel
            value[StatisticGraph.HEALTHY_COVERAGE_H] = value[StatisticGraph.NUM_OF_HEALTHY_PIXEL_H] / image_pixel
            value[StatisticGraph.BLEACHED_COVERAGE_H] = value[StatisticGraph.NUM_OF_BLEACHED_PIXEL_H] / image_pixel
            if value[StatisticGraph.NUM_OF_PIXELS_H] == 0:
                value[StatisticGraph.HEALTHY_DISTRIBUTION_H] = 0
                value[StatisticGraph.BLEACHED_DISTRIBUTION_H] = 0
            else:
                value[StatisticGraph.HEALTHY_DISTRIBUTION_H] = value[StatisticGraph.NUM_OF_HEALTHY_PIXEL_H] / value[StatisticGraph.NUM_OF_PIXELS_H]
                value[StatisticGraph.BLEACHED_DISTRIBUTION_H] = value[StatisticGraph.NUM_OF_BLEACHED_PIXEL_H] / value[StatisticGraph.NUM_OF_PIXELS_H]

        return data
            
    def export_excel(self, output_path: str):
        wb = Workbook()
        ws = wb.active
        print(f"image: {self.json_item['images']}")
        filename = self.json_item["images"][0]["filename"]
        ws.title = filename

        ws['A1'] = "Image Name"
        ws['B1'] = filename

        ws['A2'] = "Image Pixel"
        ws['B2'] = self.json_item["images"][0]["width"] * self.json_item["images"][0]["height"]

        ws["A3"] = "Export Data"
        ws["B3"] = datetime.now().strftime("%d/%m/%Y")

        headers = [
            StatisticGraph.CORAL_H,
            StatisticGraph.CORAL_ID_H,
            StatisticGraph.NUM_OF_PIXELS_H,
            StatisticGraph.NUM_OF_HEALTHY_PIXEL_H,
            StatisticGraph.NUM_OF_BLEACHED_PIXEL_H,
            StatisticGraph.CORAL_COVERAGE_H,
            StatisticGraph.HEALTHY_COVERAGE_H,
            StatisticGraph.BLEACHED_COVERAGE_H,
            StatisticGraph.HEALTHY_DISTRIBUTION_H,
            StatisticGraph.BLEACHED_DISTRIBUTION_H,
            StatisticGraph.NUM_OF_COLONY_H
        ]

        ws.append([])
        ws.append(headers)

        data = self.extract_excel_data()
        
        # sort data by coral id
        data = dict(sorted(data.items()))

        for _, value in data.items():
            print(f"value: {value}")
            row = [value[header] for header in headers]
            ws.append(row)

        ws.append([])

        # Calculate the sum of each column
        sum_info_list = []
        for header in headers:
            if header == StatisticGraph.CORAL_H:
                sum_info_list.append("")
                continue

            if header == StatisticGraph.CORAL_ID_H:
                sum_info_list.append("Sum")
                continue

            values = [value[header] for _, value in data.items()]
            sum_info_list.append(sum(values))
        ws.append(sum_info_list)

        # Calculate the average of each column
        average_info_list = []
        for header in headers:
            if header == StatisticGraph.CORAL_H:
                average_info_list.append("")
                continue

            if header == StatisticGraph.CORAL_ID_H:
                average_info_list.append("Average")
                continue

            values = [value[header] for _, value in data.items()]
            if len(data) == 0:
                average_info_list.append(0)
            else:
                average_info_list.append(sum(values) / len(data))
        ws.append(average_info_list)

        # Calculate the standard deviation of each column
        std_info_list = []
        for header in headers:
            if header == StatisticGraph.CORAL_H:
                std_info_list.append("")
                continue

            if header == StatisticGraph.CORAL_ID_H:
                std_info_list.append("STD")
                continue

            values = [value[header] for _, value in data.items()]
            if len(data) == 0:
                std_info_list.append(0)
            else:
                std_info_list.append(np.std(values))
        ws.append(std_info_list)


        ws.append([StatisticGraph.CORAL_H, "The name of the coral genus"])
        ws.append([StatisticGraph.CORAL_ID_H, "The ID of the coral genus"])
        ws.append([StatisticGraph.NUM_OF_PIXELS_H, "The number of pixels of the coral genus"])
        ws.append([StatisticGraph.NUM_OF_HEALTHY_PIXEL_H, "The number of healthy pixels of the coral genus"])
        ws.append([StatisticGraph.NUM_OF_BLEACHED_PIXEL_H, "The number of bleached pixels of the coral genus"])
        ws.append([StatisticGraph.CORAL_COVERAGE_H, "The coverage of the coral genus: number of pixels / image pixel"])
        ws.append([StatisticGraph.HEALTHY_COVERAGE_H, "The coverage of the healthy coral genus: number of healthy pixels / image pixel"])
        ws.append([StatisticGraph.BLEACHED_COVERAGE_H, "The coverage of the bleached coral genus: number of bleached pixels / image pixel"])
        ws.append([StatisticGraph.HEALTHY_DISTRIBUTION_H, "The distribution of the healthy coral genus: number of healthy pixels / number of coral pixels"])
        ws.append([StatisticGraph.BLEACHED_DISTRIBUTION_H, "The distribution of the bleached coral genus: number of bleached pixels / number of coral pixels"])
        ws.append([StatisticGraph.NUM_OF_COLONY_H, "The number of coral colony"])

        wb.save(output_path)

    def plot_coral_colony_distribution(self, output_path):
        colony_distribution = {}

        for annotation in self.json_item["annotations"]:
            supercategory_name = self.id_to_supercategory[annotation["category_id"]]
            category_id = self.name_to_id[supercategory_name]
            if category_id not in colony_distribution:
                colony_distribution[category_id] = 0
            colony_distribution[category_id] += 1

        counts = list(colony_distribution.values())
        category_ids = list(colony_distribution.keys())
        total = sum(counts)
        
        if total == 0:
            return
        
        labels = [self.id_to_name[id] for id in colony_distribution.keys()]
        colors = [self.id_to_color(id) for id in colony_distribution.keys()]

        # Function to display the value
        def value_formatter(val):
            total = sum(counts)
            value = int(round(val * total / 100))
            return f'{value}'
        
        # plt.pie(counts, labels=labels, colors=colors, autopct='%1.1f%%')
        wedges, texts, autotexts = plt.pie(counts, labels=labels, colors=colors, autopct=value_formatter)
        plt.title(f"Coral Conoly Distribution (Total: {total})")
        for idx, autotext in enumerate(autotexts):
            if category_ids[idx] == 0:
                autotext.set_color("red")
            else:
                autotext.set_color("black")
        # plt.legend(loc="upper right")
        plt.savefig(output_path)
        plt.close()
        plt.clf()
        

    def plot_coral_coverage(self, output_path):
   
        coral_area = 0
        for annotation in self.json_item["annotations"]:
            coral_area += annotation["area"]

        if coral_area == 0:
            return
        
        image_area = self.json_item["images"][0]["width"] * self.json_item["images"][0]["height"]
        non_coral_area = image_area - coral_area 

        areas = [coral_area, non_coral_area]
        labels = ["Coral", "Non-Coral"]

        plt.pie(areas, labels=labels, autopct='%1.1f%%')
        plt.title(f"Coral Coverage")
        plt.savefig(output_path)
        plt.close()
        plt.clf()

    def plot_coral_species_distribution(self, output_path):
        species_distribution = {}

        for annotation in self.json_item["annotations"]:
            supercategory_name = self.id_to_supercategory[annotation["category_id"]]
            category_id = self.name_to_id[supercategory_name]
            if category_id not in species_distribution:
                species_distribution[category_id] = 0
            species_distribution[category_id] += annotation["area"]

        if len(species_distribution) == 0:
            return
        
        areas = list(species_distribution.values())
        category_ids = list(species_distribution.keys())
        labels = [self.id_to_name[id] for id in category_ids]
        colors = [self.id_to_color(id) for id in category_ids]

        _, _, autotexts = plt.pie(areas, labels=labels, colors=colors, autopct='%1.1f%%')
        for idx, autotext in enumerate(autotexts):
            if category_ids[idx] == 0:
                autotext.set_color("red")
            else:
                autotext.set_color("black")

        plt.title(f"Coral Species Distribution")
        plt.savefig(output_path)
        plt.close()
        plt.clf()

        pass

    def plot_coral_condition_distribution(self, output_path):
        condition_distribution = {}
        for annotation in self.json_item["annotations"]:
            category_id = annotation["category_id"]
            category_name = self.id_to_name[category_id]
            condition = "Healthy"
            if self.is_category_name_is_bleached(category_name):
                condition = "Bleached"
            elif category_id == 0:
                condition = "Dead"
            
            if condition not in condition_distribution:
                condition_distribution[condition] = 0
            condition_distribution[condition] += annotation["area"]

        if len(condition_distribution) == 0:
            return

        areas = []
        for condition in ["Healthy", "Bleached", "Dead"]:
            if condition in condition_distribution:
                areas.append(condition_distribution[condition])
            else:
                areas.append(0)
        labels = ["Healthy", "Bleached", "Dead"]
        colors = ["green", "gray", "black"]

        _, _, autotexts = plt.pie(areas, labels=labels, colors=colors, autopct='%1.1f%%')
        autotexts[2].set_color("red")
        plt.title(f"Coral Condition Distribution")
        plt.savefig(output_path)
        plt.close()
        plt.clf()

        pass

    def plot_coral_species_condition_distribution(self, label_id, output_path):
        species_condition_distribution = {}

        for annotation in self.json_item["annotations"]:
            category_id = annotation["category_id"]
            category_name = self.id_to_name[category_id]
            condition = "Healthy"

            supercategory_name = self.id_to_supercategory[category_id]
            supercategory_id = self.name_to_id[supercategory_name]
            if supercategory_id != label_id:
                continue

            if self.is_category_name_is_bleached(category_name):
                condition = "Bleached"
            
            if condition not in species_condition_distribution:
                species_condition_distribution[condition] = 0
            species_condition_distribution[condition] += annotation["area"]

        if len(species_condition_distribution) == 0:
            return
        
        areas = []
        for condition in ["Healthy", "Bleached"]:
            if condition in species_condition_distribution:
                areas.append(species_condition_distribution[condition])
            else:
                areas.append(0)

        labels = ["Healthy", "Bleached"]
        colors = [self.id_to_color(label_id), "gray"]

        plt.pie(areas, labels=labels, colors=colors, autopct='%1.1f%%')
        plt.title(f"Coral Condition Distribution {self.id_to_name[label_id]}")
        plt.savefig(output_path)
        plt.close()
        plt.clf()
        
    def plot_coral_all_species_condition_distribution(self, output_path):
        for category in self.json_item["categories"]:
            category_id = category["id"]
            category_name = category["name"]
            if category_id == 0:
                continue
            elif self.is_category_name_is_bleached(category_name):
                continue
            else:
                species_output_path = os.path.join(output_path, f"species_condition_distribution_{category_name}.png")
                self.plot_coral_species_condition_distribution(category_id, species_output_path)

    def id_to_color(self, id: int) -> str:
        return self.label_colors[id]
    
    def get_name_to_id(self) -> int:
        name_to_id = {}
        for category in self.json_item["categories"]:
            name_to_id[category["name"]] = category["id"]
        return name_to_id
    
    def get_id_to_name(self) -> Dict:
        id_to_name = {}
        for category in self.json_item["categories"]:
            id_to_name[category["id"]] = category["name"]
        return id_to_name
    
    def get_id_to_supercategory(self):
        id_to_supercategory = {}
        for category in self.json_item["categories"]:
            id_to_supercategory[category["id"]] = category["supercategory"]
        return id_to_supercategory
    
    def is_category_name_is_bleached(self, category_name: str) -> bool:
        return category_name.startswith("Bleached ")


